# -*- encoding: utf-8 -*-
"""
Copyright (c) 2019 - present AppSeed.us
"""

from multiprocessing import context
from django import template
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.urls import reverse

from .models import *
import requests


#excel
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import os

def excel_upload_operations(filename,hadith_book_sub_chapter):
    print("filename:",filename)
    temp_pre_tc_obj = []
    data = dict()
    data['total_records_added'] = 0
    data['excel_max_row'] = 0
    data['excel_max_col'] = 0

    wb = load_workbook(filename)
    sheet = wb.worksheets[0]
    data['excel_max_row'] = row = sheet.max_row
    data['excel_max_col'] = col = sheet.max_column

    print("ROW :",data['excel_max_row'])
    print("COL :",data['excel_max_col'])


    listtab = []

    for r in range(1,row+1):
        listtab.append([])



    # retrive Data from Excel 
    for r in range(1,row+1):
        for c in range(1,col+1):
            e = sheet.cell(row=r,column=c)
            listtab[r-1].append(e.value)

    from .models import HadithBookContent
    total_none = 0
    total_h = 0
    r = 1
    print("listtab", listtab)
    print("listtab[r][0]", listtab[r][0])
    print("listtab[r-1][0]", listtab[r-1][0])
    for r in range(1,row):
        print(r)
        sr_no = listtab[r][0]
        arabic_content = listtab[r][1]
        roman_urdu_content = listtab[r][2]
        hindi_content = listtab[r][3]
        reference_field = listtab[r][4]
        grade = listtab[r][5]
        # print(sr_no,arabic_content,roman_urdu_content,hindi_content,reference_field,grade)
        

  
        if listtab[r][0] is None or listtab[r][0] ==  " ":
            total_none+=1
            # print(listtab[r-1][0])

        else:
            total_h+=1
            # print("else ", listtab[r-1][0])
            # print(sr_no,type(sr_no))
            # print(arabic_content, type(arabic_content))


            pre_tc_instance = HadithBookContent.objects.create(
            hadith_book_sub_chapter = hadith_book_sub_chapter,
            sr_no=listtab[r][0],
            arabic_content=listtab[r][1],
            roman_urdu_content=listtab[r][2],
            hindi_content=listtab[r][3],
            reference_field=listtab[r][4],
            grade=listtab[r][5],
            )
    print("total_none:",total_none,"total_h:",total_h)
        # if pre_tc_instance:
        #     data['total_records_added'] +=1 
        #     temp_pre_tc_obj.append(pre_tc_instance)
    return True
    return data,temp_pre_tc_obj









def quran_list_chapter(request):


    url = "https://api.quran.com/api/v4/chapters"
    payload = "{}"
    response = requests.request("GET", url, data=payload)

    json_response = response.json()
    chapter_list = json_response['chapters']
    context = {"chapter_list":chapter_list}

    html_template = loader.get_template('frontend/home.html')
    return HttpResponse(html_template.render(context, request))


def quran_get_chapter(request,chapter_id):

    url = "https://api.quran.com/api/v4/chapters/{id}?language=en".format(id=chapter_id)
    payload = "{}"
    response = requests.request("GET", url, data=payload)
    json_response = response.json()
    chapter = json_response['chapter']

    quran_obj = QuranChapterOld.objects.filter(chapter_no=chapter_id).first()

    context = {"chapter":chapter,"quran":quran_obj}
    html_template = loader.get_template('frontend/quran_get_chapter.html')
    return HttpResponse(html_template.render(context, request))


def pdf_book_list(request):
    book_list = Book.objects.prefetch_related('bookpdf_set').all()
    context = {"book_list":book_list}
    html_template = loader.get_template('frontend/pdf_book_list.html')
    return HttpResponse(html_template.render(context,request))

# mishkat-al-masabih/
def hadith_mainchapter(request):
    from .models import HadithBookMainChapter
    hadith_main_chapter_list = HadithBookMainChapter.objects.filter(hadithbook_id=1)
    print(hadith_main_chapter_list)
    context = {"hadith_main_chapter_list":hadith_main_chapter_list}
    html_template = loader.get_template('frontend/hadees_mainchapter.html')
    return HttpResponse(html_template.render(context,request))

# mishkat-al-masabih/faith/ XX
# mishkat-al-masabih/id/ 
def hadith_subchapter(request,main_chp):
    from .models import HadithBookMainChapter,HadithBookSubChapter
    print("$$$ main_chp $$$", main_chp)
    obj = HadithBookMainChapter.objects.filter(english_name=main_chp).first()
    hadith_sub_chapter = HadithBookSubChapter.objects.filter(hadith_book_main_chapter_id=obj.id)

    context = {"hadith_sub_chapter":hadith_sub_chapter}
    html_template = loader.get_template('frontend/hadees_subchapter.html')
    return HttpResponse(html_template.render(context,request))

# mishkat-al-masabih/faith/
def hadith_content(request,main_chp,sub_chp_id,sr_no):
    try:
        print("share_hadith",main_chp)
        # main_chp = HadithBookMainChapter.objects.filter(english_name=main_chp).first()
        hadith_sub_chapter = HadithBookContent.objects.filter(hadith_book_sub_chapter_id=sub_chp_id,sr_no=sr_no).first()
        print("sub_chapter_name",hadith_sub_chapter.hadith_book_sub_chapter)
        context = {"isfound":True,"content":hadith_sub_chapter,"hadith_book_main_chapter":main_chp,"sub_chapter_name":hadith_sub_chapter.hadith_book_sub_chapter.sub_chapter_name}
        html_template = loader.get_template('frontend/hadith_single_content.html')
        return HttpResponse(html_template.render(context,request))
    except:

        context = {"isfound":False}
        html_template = loader.get_template('frontend/hadith_single_content.html')
        return HttpResponse(html_template.render(context,request))

def t_hadith_content(request,main_chp,id_no):
    print("****sr_no",id_no,main_chp)
    try:
        print("share_hadith",main_chp)
        # main_chp = HadithBookMainChapter.objects.filter(english_name=main_chp).first()
        hadith_sub_chapter = HadithBookContent.objects.filter(id=id_no).first()
        print("sub_chapter_name",hadith_sub_chapter.hadith_book_sub_chapter)
        context = {"isfound":True,"content":hadith_sub_chapter,"hadith_book_main_chapter":main_chp,"sub_chapter_name":hadith_sub_chapter.hadith_book_sub_chapter.sub_chapter_name}
        html_template = loader.get_template('frontend/hadith_single_content.html')
        return HttpResponse(html_template.render(context,request))
    except:

        context = {"isfound":False}
        html_template = loader.get_template('frontend/hadith_single_content.html')
        return HttpResponse(html_template.render(context,request))



def hadith_book_excel_file(request):
    context = {"isfound":False}
    html_template = loader.get_template('frontend/hadith_single_content.html')
    return HttpResponse(html_template.render(context,request))





@login_required(login_url="/login/")
def index(request):
    context = {'segment': 'index'}

    html_template = loader.get_template('home/index.html')
    return HttpResponse(html_template.render(context, request))


@login_required(login_url="/login/")
def pages(request):
    context = {}
    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:

        load_template = request.path.split('/')[-1]

        if load_template == 'admin':
            return HttpResponseRedirect(reverse('admin:index'))
        context['segment'] = load_template

        html_template = loader.get_template('home/' + load_template)
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:

        html_template = loader.get_template('home/page-404.html')
        return HttpResponse(html_template.render(context, request))

    except:
        html_template = loader.get_template('home/page-500.html')
        return HttpResponse(html_template.render(context, request))


