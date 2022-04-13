# -*- encoding: utf-8 -*-
"""
Copyright (c) 2019 - present AppSeed.us
"""

from operator import mod
from pyexpat import model
from statistics import mode
from turtle import title, update
from venv import create
from django.db import models
from django.contrib.auth.models import User
from django.template.defaultfilters import slugify
from django.urls import reverse


# from .views import excel_upload_operations

from uuid import uuid4

def generateUUID():
    return str(uuid4())

from tinymce.models import HTMLField


# Create your models here.
def BookPDF_upload_location(instance, filename):
    model_name = instance.hadith_book_main_chapter.title.lower().replace(" ", "-")
    file_name = filename.lower().replace(" ", "-")
    return "book/{}/{}".format(model_name, file_name)

def hadith_excel_upload_location(instance, filename):
    model_name = instance.hadith_book_main_chapter.english_name.lower().replace(" ", "-")
    file_name = filename.lower().replace(" ", "-")
    return "Hadith/{}/{}".format(model_name, file_name)


class Book(models.Model):
    title = models.TextField(blank=False,null=False)
    volumes = models.TextField(blank=True,null=True)
    description = models.TextField(blank=True,null=True)
    
    def __str__(self):
        return self.title

class BookPDF(models.Model):
    # book = models.ForeignKey(Book, on_delete=models.SET_NULL, null=True, blank=True, related_name='book')
    book = models.ForeignKey(Book,on_delete=models.SET_NULL,null=True, blank=True,)
    title = models.TextField(blank=False,null=False)
    total_pages = models.TextField(blank=True,null=True)
    file_url = models.FileField(upload_to=BookPDF_upload_location, null = True, blank = True)
    external_file_url = models.TextField(blank=True,null=True)
    
    def __str__(self):
        return self.title


class QuranChapterOld(models.Model):
    chapter_no = models.IntegerField(null=False,blank=False)
    chapter_content = HTMLField(null=False,blank=False)
    is_visible = models.BooleanField(default=False)
    created_at = models.DateTimeField(auto_now_add=True)
    updated_at = models.DateTimeField(auto_now=True)

    def __str__(self):
        return str(self.chapter_no)

    def get_absolute_url(self):
        return reverse("quran_get_chapter", kwargs={"chapter_id":self.id})


#Mishkatul Masabeeh
class HadithBook(models.Model):
    title = models.CharField(max_length=100,null=False,blank=False)
    description = models.TextField(null=True,blank=True)
    slug = models.SlugField(unique=False)

    def __str__(self):
        return self.title

    def save(self, *args, **kwargs):
        self.slug = self.slug or slugify(self.title)
        super().save(*args, **kwargs)


#Mishkatul Masabeeh #Faith (main chapter)
class HadithBookMainChapter(models.Model):
    hadithbook = models.ForeignKey(HadithBook,on_delete=models.SET_NULL,null=True)
    chapter_no = models.IntegerField(max_length=3)
    english_name = models.CharField(max_length=200)
    arabic_name = models.CharField(max_length=200)

    def __str__(self):
        return self.english_name

    def get_absolute_url(self):
        return reverse("hadith_subchapter", kwargs={"main_chp":self.english_name})


#Mishkatul Masabeeh -> #Faith -> #sub chapters
class HadithBookSubChapter(models.Model):
    hadith_book_main_chapter = models.ForeignKey(HadithBookMainChapter,on_delete=models.SET_NULL,null=True)
    sub_chapter_name = models.CharField(max_length=200)

    def __str__(self):
        return f"{self.hadith_book_main_chapter} - {self.sub_chapter_name}"

    def get_absolute_url(self):
        return reverse("hadith_subchapter", kwargs={"main_chp":self.hadith_book_main_chapter.english_name.english_name})

HADITHGRADE = [
    ("Muttafaqun 'alayh","Muttafaqun 'alayh"),
    ("Sahih","Sahih"),
    ("Hasan","Hasan"),
    ("Zaeef","Zaeef"),
]



#Mishkatul Masabeeh -> #Faith -> #sub chapters -> content
class HadithBookContent(models.Model):
    hadith_book_sub_chapter = models.ForeignKey(HadithBookSubChapter,on_delete=models.SET_NULL,null=True)
    sr_no = models.IntegerField(max_length=4,null=True,blank=True)
    arabic_content = models.TextField(null=True,blank=True)
    roman_urdu_content = models.TextField(null=True,blank=True)
    hindi_content = models.TextField(null=True,blank=True)
    reference_field = models.TextField(null=True,blank=True)
    grade = models.CharField(choices=HADITHGRADE,default="Sahih", max_length=50,null=True,blank=True)

    def __str__(self):
        return f"{self.hadith_book_sub_chapter}"

    def get_absolute_url(self):
        print("self.hadith_book_sub_chapter.id", self.hadith_book_sub_chapter.sub_chapter_name)
        print("main chapter", self.hadith_book_sub_chapter.hadith_book_main_chapter.hadithbook.title)
        return reverse("t_hadith_content", kwargs={"main_chp":self.hadith_book_sub_chapter.hadith_book_main_chapter.english_name,"id_no":self.id})

        # return reverse("hadith_content", kwargs={"main_chp":self.hadith_book_sub_chapter.hadith_book_main_chapter.hadithbook.title,"sub_chp_id": self.hadith_book_sub_chapter.id,"sr_no":self.sr_no})
        # return f"mishkat-al-masabih/Faith/1/1/"

from smart_selects.db_fields import ChainedForeignKey

class HadithBookExcelFile(models.Model):
    file_url = models.FileField(upload_to=hadith_excel_upload_location,null=False,blank=False)
    Hadithbook = models.ForeignKey(HadithBook,on_delete=models.SET_NULL,null=True,related_name="Hadithbook_excel")
    hadith_book_main_chapter = ChainedForeignKey(
        HadithBookMainChapter,
        chained_field="Hadithbook",
        chained_model_field="hadithbook",
        show_all=False,
        auto_choose=True,
        sort=True,
        related_name="hadith_book_main_chapter_excel")

    Hadith_book_sub_chapter = ChainedForeignKey(
        HadithBookSubChapter,
        chained_field="hadith_book_main_chapter",
        chained_model_field="hadith_book_main_chapter",
        show_all=False,
        auto_choose=True,
        sort=True)

    def __str__(self):
        return str(self.Hadithbook)

    def save(self, *args, **kwargs):
        super(HadithBookExcelFile, self).save(*args, **kwargs)
        filename = self.file_url
        hadith_book_sub_chapter = self.Hadith_book_sub_chapter
        excel_upload_operations(filename,hadith_book_sub_chapter)

        # data,temp_pre_tc_obj = excel_upload_operations(filename,hadith_book_sub_chapter)
        # print(data)
        # print(temp_pre_tc_obj)
#excel
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import os

def excel_upload_operations(filename,hadith_book_sub_chapter):
    print("filename:",filename)
    temp_pre_tc_obj = []
    data = dict()
    data['total_records_added'] = 0
    data['excel_max_row'] = 0
    data['excel_max_col'] = 0

    wb = load_workbook(filename)
    sheet = wb.worksheets[0]
    data['excel_max_row'] = row = sheet.max_row
    data['excel_max_col'] = col = sheet.max_column

    print("ROW :",data['excel_max_row'])
    print("COL :",data['excel_max_col'])


    listtab = []

    for r in range(1,row+1):
        listtab.append([])



    # retrive Data from Excel 
    for r in range(1,row+1):
        for c in range(1,col+1):
            e = sheet.cell(row=r,column=c)
            listtab[r-1].append(e.value)

    from .models import HadithBookContent
    total_none = 0
    total_h = 0
    r = 1
    print("listtab", listtab)
    print("listtab[r][0]", listtab[r][0])
    print("listtab[r-1][0]", listtab[r-1][0])
    for r in range(1,row):
        print(r)
        sr_no = listtab[r][0]
        arabic_content = listtab[r][1]
        roman_urdu_content = listtab[r][2]
        hindi_content = listtab[r][3]
        reference_field = listtab[r][4]
        grade = listtab[r][5]
        # print(sr_no,arabic_content,roman_urdu_content,hindi_content,reference_field,grade)
        

  
        if listtab[r][0] is None or listtab[r][0] ==  " ":
            total_none+=1
            # print(listtab[r-1][0])

        else:
            total_h+=1
            # print("else ", listtab[r-1][0])
            # print(sr_no,type(sr_no))
            # print(arabic_content, type(arabic_content))


            pre_tc_instance = HadithBookContent.objects.create(
            hadith_book_sub_chapter = hadith_book_sub_chapter,
            sr_no=listtab[r][0],
            arabic_content=listtab[r][1],
            roman_urdu_content=listtab[r][2],
            hindi_content=listtab[r][3],
            reference_field=listtab[r][4],
            grade=listtab[r][5],
            )
    print("total_none:",total_none,"total_h:",total_h)
        # if pre_tc_instance:
        #     data['total_records_added'] +=1 
        #     temp_pre_tc_obj.append(pre_tc_instance)
    return True
    return data,temp_pre_tc_obj